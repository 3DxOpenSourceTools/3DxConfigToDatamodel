import xml.etree.ElementTree as ET
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox


def parse_xml(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    attribute_defs = []

    for parameter in root.findall('Parameter'):
        if parameter.get('category') == 'AttributeDef':
            value = parameter.text
            parts = value.split('|')
            if len(parts) >= 3:
                class_name = parts[0]
                attr_type = parts[1]
                property_name = parts[2]
                attribute_defs.append((class_name, attr_type, property_name))

    return attribute_defs


def add_to_excel(attribute_defs, excel_path):
    if os.path.exists(excel_path):
        workbook = openpyxl.load_workbook(excel_path)
    else:
        workbook = openpyxl.Workbook()

    if 'classes' not in workbook.sheetnames:
        class_sheet = workbook.create_sheet(title='classes')
        class_sheet.append(["Class", "Attribute", "Type", "Enum"])
    else:
        class_sheet = workbook['classes']

    for class_name, attr_type, property_name in attribute_defs:
        class_sheet.append(['XP_' + class_name + 'Ext_',
                           property_name, attr_type, ""])

    workbook.save(excel_path)
    print(f"Data added to {excel_path}")


def main():
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    xml_file_path = filedialog.askopenfilename(
        title="Select XML file",
        filetypes=(("XML files", "*.xml"), ("All files", "*.*"))
    )

    if not xml_file_path:
        print("No XML file selected.")
        return

    attribute_defs = parse_xml(xml_file_path)

    if not attribute_defs:
        print("No AttributeDef found in the XML file.")
        return

    response = messagebox.askyesno("Add to existing configuration",
                                   "Do you want to add the configuration to an existing Excel file?")

    if response:
        excel_file_path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if not excel_file_path:
            print("No Excel file selected.")
            return
    else:
        excel_file_path = filedialog.asksaveasfilename(
            title="Save Excel file",
            defaultextension=".xlsx",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if not excel_file_path:
            print("No Excel file selected.")
            return

    add_to_excel(attribute_defs, excel_file_path)


if __name__ == "__main__":
    main()
