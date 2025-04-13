import openpyxl
import csv
import tkinter as tk
from tkinter import filedialog


def parse_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)

    interfaces = set()
    realizations = set()

    # Parse interfaces sheet
    if "interfaces" in workbook.sheetnames:
        interface_sheet = workbook["interfaces"]
        for row in interface_sheet.iter_rows(min_row=2, values_only=True):
            interface_name = row[0]
            if interface_name:
                interfaces.add(interface_name)

    # Parse realizations sheet
    if "realisation" in workbook.sheetnames:
        realization_sheet = workbook["realisation"]
        for row in realization_sheet.iter_rows(min_row=2, values_only=True):
            source, target = row
            if target:
                realizations.add(target)

    return interfaces, realizations


def list_interfaces_without_realization(interfaces, realizations):
    interfaces_without_realization = interfaces - realizations
    return interfaces_without_realization


def save_to_csv(interfaces_without_realization, output_file):
    with open(output_file, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Interface"])
        for interface in interfaces_without_realization:
            writer.writerow([interface])
    print(f"CSV report saved to {output_file}")


def main():
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    excel_file_path = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )

    if not excel_file_path:
        print("No Excel file selected.")
        return

    interfaces, realizations = parse_xlsx(excel_file_path)
    interfaces_without_realization = list_interfaces_without_realization(
        interfaces, realizations)

    if interfaces_without_realization:
        print("Interfaces without realization:")
        for interface in interfaces_without_realization:
            print(interface)

        output_file_path = filedialog.asksaveasfilename(
            title="Save CSV report",
            defaultextension=".csv",
            filetypes=(("CSV files", "*.csv"), ("All files", "*.*"))
        )

        if not output_file_path:
            print("No output file selected.")
            return

        save_to_csv(interfaces_without_realization, output_file_path)
    else:
        print("All interfaces have realizations.")


if __name__ == "__main__":
    main()
