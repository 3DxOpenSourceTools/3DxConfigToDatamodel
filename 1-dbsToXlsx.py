import json
import os
import tkinter as tk
from tkinter import filedialog
import openpyxl


def find_types(data):
    types = {}
    interfaces = {}
    for key, value in data.items():
        if isinstance(value, dict):
            if 'Types' in value or 'Interfaces' in value:
                types = value.get('Types', {})
                interfaces = value.get('Interfaces', {})
                return types, interfaces
            else:
                result_types, result_interfaces = find_types(value)
                if result_types or result_interfaces:
                    return result_types, result_interfaces
    return types, interfaces


def parse_json(file_path):
    with open(file_path, 'r') as file:
        data = json.load(file)

    types, interfaces = find_types(data)
    return types, interfaces


def clean_name(name):
    return name.split('__')[0]


def export_enums_to_xlsx(enums, classes, interfaces, generalizations, realisations, output_dir):
    workbook = openpyxl.Workbook()

    # Create enum sheet
    enum_sheet = workbook.active
    enum_sheet.title = "enum"

    # Add headers to enum sheet
    enum_sheet.append(["enum", "properties"])

    # Add enum values to enum sheet
    for enum_name, values in enums.items():
        for value in values:
            enum_sheet.append([enum_name, value])

    # Create classes sheet
    class_sheet = workbook.create_sheet(title="classes")

    # Add headers to classes sheet
    class_sheet.append(["Class", "Attribute", "Type", "Enum"])

    # Add class attributes to classes sheet
    for class_name, attributes in classes.items():
        if not attributes:
            class_sheet.append([class_name, "", "", ""])
        for attr_name, attr_type, attr_enum in attributes:
            class_sheet.append([class_name, attr_name, attr_type, attr_enum])

    # Create interfaces sheet
    interface_sheet = workbook.create_sheet(title="interfaces")

    # Add headers to interfaces sheet
    interface_sheet.append(["Interface", "Attribute", "Type"])

    # Add interface attributes to interfaces sheet
    for interface_name, attributes in interfaces.items():
        if not attributes:
            interface_sheet.append([interface_name, "", ""])
        for attr_name, attr_type in attributes:
            interface_sheet.append([interface_name, attr_name, attr_type])

    # Create generalizations sheet
    generalization_sheet = workbook.create_sheet(title="generalization")

    # Add headers to generalization sheet
    generalization_sheet.append(
        ["sourceGeneralization", "targetGeneralization"])

    # Add generalization values to generalization sheet
    for source, target in generalizations.items():
        generalization_sheet.append([source, target])

    # Create realisations sheet
    realisation_sheet = workbook.create_sheet(title="realisation")

    # Add headers to realisation sheet
    realisation_sheet.append(["sourceRealisation", "targetRealisation"])

    # Add realisation values to realisation sheet
    for source, targets in realisations.items():
        for target in targets:
            realisation_sheet.append([source, target])

    # Save the workbook
    output_file = os.path.join(
        output_dir, "enums_classes_interfaces_generalizations_realisations.xlsx")
    workbook.save(output_file)
    print(
        f"Enumerations, classes, interfaces, generalizations, and realisations exported to {output_file}")


def main():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(
        title="Select JSON file",
        filetypes=(("JSON files", "*.json"), ("All files", "*.*"))
    )

    if not file_path:
        print("No file selected.")
        return

    types, interfaces = parse_json(file_path)

    enums = {}
    classes = {}
    interface_attributes = {}
    generalizations = {}
    realisations = {}

    # Process Types
    for type_name, type_info in types.items():
        class_name = type_info.get('name', type_name)
        attributes = type_info.get('Attributes', {})
        class_attributes = []
        for attr_name, attr_info in attributes.items():
            attr_display_name = attr_info.get('name', attr_name)
            attr_type = attr_info.get('Type', 'attribute')
            attr_enum = "enum" if 'AuthorizedValues' in attr_info else "attribute"
            class_attributes.append((attr_display_name, attr_type, attr_enum))
            authorized_values = attr_info.get('AuthorizedValues')
            if authorized_values:
                enum_name = f"{class_name}_{attr_name}_Enum"
                enums[enum_name] = authorized_values
        classes[class_name] = class_attributes

        # Check for Parent and add to generalizations
        parent = type_info.get('Parent')
        if parent:
            generalizations[class_name] = parent

    # Process Interfaces
    for interface_name, interface_info in interfaces.items():
        class_name = interface_info.get('name', interface_name)
        attributes = interface_info.get('Attributes', {})
        interface_attrs = []
        for attr_name, attr_info in attributes.items():
            attr_display_name = attr_info.get('name', attr_name)
            attr_type = attr_info.get('Type', 'string')
            interface_attrs.append((attr_display_name, attr_type))
            authorized_values = attr_info.get('AuthorizedValues')
            if authorized_values:
                enum_name = f"{class_name}_{attr_name}_Enum"
                enums[enum_name] = authorized_values
        interface_attributes[class_name] = interface_attrs

        # Check for ScopeTypes and add to realisations
        scope_types = interface_info.get('ScopeTypes', [])
        if scope_types:
            realisations[class_name] = scope_types

    output_dir = filedialog.askdirectory(title="Select Output Directory")
    if not output_dir:
        print("No output directory selected.")
        return

    export_enums_to_xlsx(enums, classes, interface_attributes,
                         generalizations, realisations, output_dir)


if __name__ == "__main__":
    main()
