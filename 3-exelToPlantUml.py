import openpyxl
import os
import tkinter as tk
from tkinter import filedialog
from plantuml import PlantUML


def parse_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)

    enums = {}
    classes = {}
    interfaces = {}
    generalizations = []
    realisations = []

    # Parse enum sheet
    if "enum" in workbook.sheetnames:
        enum_sheet = workbook["enum"]
        for row in enum_sheet.iter_rows(min_row=2, values_only=True):
            enum_name, value = row
            if enum_name not in enums:
                enums[enum_name] = []
            enums[enum_name].append(value)

    # Parse classes sheet
    if "classes" in workbook.sheetnames:
        class_sheet = workbook["classes"]
        for row in class_sheet.iter_rows(min_row=2, values_only=True):
            class_name, attr_name, attr_type, attr_enum = row
            if class_name not in classes:
                classes[class_name] = []
            classes[class_name].append((attr_name, attr_type, attr_enum))

    # Parse interfaces sheet
    if "interfaces" in workbook.sheetnames:
        interface_sheet = workbook["interfaces"]
        for row in interface_sheet.iter_rows(min_row=2, values_only=True):
            interface_name, attr_name, attr_type = row
            if interface_name not in interfaces:
                interfaces[interface_name] = []
            interfaces[interface_name].append((attr_name, attr_type))

    # Parse generalization sheet
    if "generalization" in workbook.sheetnames:
        generalization_sheet = workbook["generalization"]
        for row in generalization_sheet.iter_rows(min_row=2, values_only=True):
            source, target = row
            generalizations.append((source, target))

    # Parse realisation sheet
    if "realisation" in workbook.sheetnames:
        realisation_sheet = workbook["realisation"]
        for row in realisation_sheet.iter_rows(min_row=2, values_only=True):
            source, target = row
            realisations.append((source, target))

    return enums, classes, interfaces, generalizations, realisations


def generate_class_diagram(class_name, enums, classes, interfaces, generalizations, realisations, output_dir):
    uml_code = "@startuml\n\n"

    # Add the class and its attributes
    if class_name in classes:
        uml_code += f"class {class_name} {{\n"
        for attr_name, attr_type, attr_enum in classes[class_name]:
            if attr_enum == "enum":
                uml_code += f"  {attr_name}: {class_name}_{attr_name}_Enum\n"
            else:
                uml_code += f"  {attr_name}: {attr_type}\n"
        uml_code += "}\n\n"

    # Add generalizations
    added_classes = set()  # Track added classes

    for source, target in generalizations:
        if source == class_name or target == class_name:
            uml_code += f"{source} --|> {target}\n"
            if target in classes and target not in added_classes:
                uml_code += f"\nclass {target} {{\n"
                for attr_name, attr_type, attr_enum in classes[target]:
                    uml_code += f"  {attr_name}: {attr_type}\n"
                uml_code += "}\n\n"
                added_classes.add(target)  # Mark the class as added

    # Add realisations
    for source, target in realisations:
        if source == class_name or target == class_name:
            uml_code += f"{target} ..|>{source} \n"
            if source in interfaces:
                uml_code += f"\nclass {source} <<interface>> {{\n"
                for attr_name, attr_type in interfaces[source]:
                    uml_code += f"  {attr_name}: {attr_type}\n"
                uml_code += "}\n\n"

    # Add linked enums
    for attr_name, attr_type, attr_enum in classes[class_name]:
        if attr_enum == "enum":
            enum_name = f"{class_name}_{attr_name}_Enum"
            if enum_name in enums:
                uml_code += f"enum {enum_name} {{\n"
                for value in enums[enum_name]:
                    uml_code += f"  {value}\n"
                uml_code += "}\n\n"

    uml_code += "@enduml"

    output_file = os.path.join(output_dir, f"{class_name}_diagram.puml")
    with open(output_file, 'w') as file:
        file.write(uml_code)

    print(f"UML diagram generated: {output_file}")


def main():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )

    if not file_path:
        print("No file selected.")
        return

    enums, classes, interfaces, generalizations, realisations = parse_xlsx(
        file_path)

    output_dir = filedialog.askdirectory(title="Select Output Directory")
    if not output_dir:
        print("No output directory selected.")
        return

    for class_name in classes.keys():
        generate_class_diagram(class_name, enums, classes,
                               interfaces, generalizations, realisations, output_dir)

    # Convert to PNG using PlantUML
    plantuml = PlantUML(url='http://www.plantuml.com/plantuml/img/')
    for class_name in classes.keys():
        puml_file = os.path.join(output_dir, f"{class_name}_diagram.puml")
        plantuml.processes_file(puml_file)
        print(
            f"UML diagram converted to PNG: {puml_file.replace('.puml', '.png')}")


if __name__ == "__main__":
    main()
